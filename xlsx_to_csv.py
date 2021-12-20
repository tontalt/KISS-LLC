#====================================================================
#
# This module uses openpyxl, xlrd
# 
#====================================================================

import openpyxl
import datetime
import dateutil
import xlrd
import sys
import math
import io
from hvmg_libx import *
from openpyxl import load_workbook

TYPE_UNKNOWN    = 0
TYPE_MARRIOTT   = 1
TYPE_HARDROCK   = 2

IN_PATH_XLSX = "F:\\temp\\input\\"
OUT_PATH_CSV = "F:\\temp\\output\\"

MARRIOTT_INPUT_FILE = "Marriott_Responses_Export_11_08_2021"
COLS_MARRIOTT_IN    = [ 2, 3, 6, 20, 22, 23, 24, 25, 26 ]
COLS_MARRIOT_OUT    = [ 1, 2, 3, 4, 5, 6, 7, 8, 9 ]

#16=unit
HARDROCK_INPUT_FILE = "Daytona Beach 2020 Export_2021-08-01_235959"
COLS_HARDROCK_IN    = [ 1, 2, 16, 35, 39, 41, 44 ]
COLS_HARDROCK_OUT   = [ 2, 1, 3, 4, 5, 7, 6 ]

#-----------------------------------------------------------------

input_file = ""
selected_cols_in = ""
selected_cols_out = ""

def get_run_type(runtype):
    try:
        if runtype == TYPE_MARRIOTT:
            input_file = MARRIOTT_INPUT_FILE
            selected_cols_in = COLS_MARRIOTT_IN
            selected_cols_out = COLS_MARRIOT_OUT
            return True
        elif runtype == TYPE_HARDROCK:
            input_file = HARDROCK_INPUT_FILE
            selected_cols_in = COLS_HARDROCK_IN
            selected_cols_out = COLS_HARDROCK_OUT
            return True
        else:
            return False
    except Exception as e:
        print("Invalide run type=" + runtype)
        exit(1)

#---------------------------------------

def sort_output_cols(list_in, list_sort):
    list_out = []
    clean_list = []

    try:
        list_out = [b"!"] * 128
        #for ii in list_sort:
        for ii in range(len(list_in)):
            list_out[list_sort[ii]] = list_in[ii]
        # clean up the extra "!"
        for ii in range(len(list_out)):
            if list_out[ii] != b"!":
                clean_list.append(list_out[ii])
    except Exception as e:
        return None
    return clean_list

#---------------------------------------

def xstr(s):
    try:
        if s is None:
            return ""
        if s == 'None':
            return ""
        if s.isascii():
            return str(s)
        else:
            return "***"
    except Exception as e:
        return ""
    return ""

#---------------------------------------

def is_null(ss):
    try:
        if ss is None:
            return True
        if ss == 'None':
            return True
    except Exception as e:
        return True
    return False

#====================================================================
#
# BEGIN mainline:
#
#====================================================================
#DEBUG  
selected_input = TYPE_MARRIOTT

if get_run_type(selected_input) != True:
    print("Invalid run type: " + selected_input)
    exit(1)

#====================================================================
# input and output filespecs
output_file     = input_file + ".csv"
input_file      = input_file + ".xlsx"
input_file      = IN_PATH_XLSX + input_file
output_file     = OUT_PATH_CSV + output_file

# input and output workbook objects
try:
    xls_wb = load_workbook(filename = input_file, data_only = True)
    csv_handle = io.FileIO(output_file, "w")
    csv_file = io.BufferedWriter(csv_handle, buffer_size=256000)    
except Exception as e:
    print("Exception: " + str(e))
    exit(1)

#====================================================================
# output data to provide one row per installment

rowcnt = 0
colcnt = 0
linelist = []
input_sheet = xls_wb.active

try:
    #Iterate through worksheet and print cell contents
    for row in input_sheet.iter_rows():
        rowcnt += 1
        colcnt = 0
        tstr = ""
        blank_line = True
        linelist.clear()
        if rowcnt % 100 == 0:
            print("Row=" + str(rowcnt))
        for ii in range(input_sheet.max_column):
            #debug
            colcnt += 1
            if rowcnt <= 2:
                continue
            if rowcnt == 3:
                brak = 1
            if rowcnt == 4:
                brak = 2
            #end debug
            mycell = input_sheet.cell(row=rowcnt, column=colcnt)
            mystr = str(mycell.value)
            if colcnt in selected_cols_in:
                #print("Cell Value=\"" + str(mycell.value) + "\"")
                if is_null(mystr):
                    linelist.append(csvnum(mystr))
                    blank_line = False
                    continue
                if is_number(mycell.value):
                    ss = xstr(mystr)
                    ss = csvnum(ss)
                    linelist.append(ss)
                else:
                    tstr = xstr(mystr)
                    linelist.append(csvstr(tstr.strip(), ','))
                blank_line = False
                # end if is_integer
            # end if colcnt in hardrock
        # end for cell in row
        if blank_line == False:
            lines_out = sort_output_cols(linelist, selected_cols_out)
            for my_bytes in lines_out:
                csv_file.write(my_bytes)
            csv_file.write("\n".encode(encoding="utf-8", errors="ignore"))
    # end for row in sheet
except Exception as e:
    print(str(e))

print("RowCnt=" + str(rowcnt))

#====================================================================
# save spreadsheet
try:
    csv_file.flush()
    csv_file.close()
    csv_handle.close()
except:
    print("Save Failed")

#====================================================================
# exit
xls_wb.close
print("Done")
